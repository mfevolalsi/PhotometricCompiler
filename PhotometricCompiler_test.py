
from pathlib import Path
import json
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date

class MeterData:
    def __init__(self,data,origin,offset,ws):
        self.data = data
        self.origin = (origin[0]+offset[0], origin[1]+offset[1])
        self.ws = ws
    
    def fill_col(self):
        for i in range(len(self.data)): self.ws.cell(row=self.origin[0]+i, column=self.origin[1], value=self.data[i])
    
    def title(self,title, title_offset):
        title_coord = ws.cell(row=coordinates[0]+offset[0], column=coordinates[1]+offset[1]).coordinate
        ws[title_coord].font = ws[title_coord].font.copy(bold=True)
        ws.cell(row=coordinates[0]+offset[0], column=coordinates[1]+offset[1], value=title)
        
    def coordinates(self):
        return self.ws.cell(row=self.origin[0], column=self.origin[1]).coordinate
    
def nearest_space(ws):
   origin = [3, 2] # [row, column]
   while ws[get_column_letter(origin[1])+str(origin[0])].value != None: origin[1] += 12
   return origin        

x = []
y = []
u = []
v = []
dev = []
cct = []

user = input("Computer Username: ")
filename = input("filename: ")
n = input("Number of fixtures: ")

## Read  values from the Photometric saved JSON
for i in range(int(n)):
    path = Path("C:\\Users\\" + user + "\\Desktop\\Testing\\" + filename + "-" + str(i+1).zfill(4))  # current directory
    extension = ".json"

    measurementData = next(path.glob(f"*{extension}"))  # returns the file with extension or None
    if measurementData:
        with open(measurementData) as file:
            data = json.load(file)
         
            # Iterating through the json
            # list
            x.append(data['CIEx'])
            y.append(data['CIEy'])
            u.append(data['CIEu'])
            v.append(data['CIEv'])
            cct.append(data['CCT'])
            
            # Closing file
            file.close()     

path = Path("C:\\Users\\" + user + "\\Desktop\\temp.xlsx")
wb = load_workbook(path)
# ws = wb.create_sheet('test')
# ws.title = 'test'

print(wb.sheetnames)
sheet = input("Type an available sheet name, or a new sheet name to create a new sheet: ")

if sheet in wb.sheetnames:
    ws = wb.get_sheet_by_name(sheet)
    # condition = input("Fixture conditions: ")
else:
    ws = wb.create_sheet(sheet)
    # today = date.today()
    # d1 = today.strftime("%m/$d/%y")
    # fix = input("Part number: Job Number: ")
    # LED = input("LED Manufacturer: ")

coordinates = nearest_space(ws)

X_data = MeterData(x,coordinates,[0,0],ws)
X_data.fill_col()
Y_data = MeterData(y,coordinates,[0,1],ws)
Y_data.fill_col()
U_data = MeterData(u,coordinates,[0,2],ws)
U_data.fill_col()
V_data = MeterData(v,coordinates,[0,3],ws)
V_data.fill_col()
CCT_data = MeterData(cct,coordinates,[0,4],ws)
CCT_data.fill_col()

Avg = MeterData(['=AVERAGE(%s:%s)' % (U_data.coordinates()[1],U_data.coordinates()[1])],coordinates,[0,7],ws)
Avg.fill_col()
# u_col = get_column_letter(coordinates[1] + 2)
# v_col = get_column_letter(coordinates[1] + 3)
# dev_col = get_column_letter(coordinates[1] + 4)

# ws.cell(row=coordinates[0], column=coordinates[1]+7, value='=AVERAGE(%s:%s)' % (u_col,u_col))
# u_avg = ws.cell(row=coordinates[0], column=coordinates[1]+7).coordinate
# ws.cell(row=coordinates[0]+1, column=coordinates[1]+7, value='=_xlfn.STDEV.P(%s:%s)' % (u_col,u_col))
# u_std = ws.cell(row=coordinates[0]+1, column=coordinates[1]+7).coordinate

# ws.cell(row=coordinates[0], column=coordinates[1]+8, value='=AVERAGE(%s:%s)' % (v_col,v_col))
# v_avg = ws.cell(row=coordinates[0], column=coordinates[1]+8).coordinate
# ws.cell(row=coordinates[0]+1, column=coordinates[1]+8, value='=_xlfn.STDEV.P(%s:%s)' % (v_col,v_col))
# v_std = ws.cell(row=coordinates[0]+1, column=coordinates[1]+8).coordinate

# for i in range(int(n)):
    # u_coor=u_col+str(coordinates[0]+i)
    # v_coor=v_col+str(coordinates[0]+i)
    # dev.append('=SQRT((%s-%s)^2+(%s-%s)^2)' % (u_coor,u_avg,v_coor,v_avg))

# fill_row(ws, coordinates, 4, dev)

# ws.cell(row=coordinates[0]+4, column=coordinates[1]+7, value='=MAX(%s:%s)/0.0011' % (dev_col,dev_col))  

# ws.cell(row=coordinates[0]+5, column=coordinates[1]+7, value='=0.0022/%s' % (u_std))  
# u_Z = ws.cell(row=coordinates[0]+5, column=coordinates[1]+7).coordinate
# ws.cell(row=coordinates[0]+5, column=coordinates[1]+8, value='=0.0022/%s' % (v_std))
# v_Z = ws.cell(row=coordinates[0]+5, column=coordinates[1]+8).coordinate

# ws.cell(row=coordinates[0]+6, column=coordinates[1]+7, value='=(1-_xlfn.NORM.S.DIST(%s,TRUE))*1000+(1-_xlfn.NORM.S.DIST(%s,TRUE))*1000' % (u_Z, v_Z))
# ws.cell(row=coordinates[0]+7, column=coordinates[1]+7, value='=(1-_xlfn.NORM.S.DIST(%s,TRUE))+(1-_xlfn.NORM.S.DIST(%s,TRUE))' % (u_Z, v_Z))

# title(ws, coordinates, offset=[-1, 0], title='X')
# title(ws, coordinates, offset=[-1, 1], title='Y')
# title(ws, coordinates, offset=[-1, 2], title='U\'')
# title(ws, coordinates, offset=[-1, 3], title='V\'')
# title(ws, coordinates, offset=[-1, 4], title='Deviation')
# title(ws, coordinates, offset=[-1, 5], title='CCT')
# title(ws, coordinates, offset=[-1, 7], title='U\'')
# title(ws, coordinates, offset=[-1, 8], title='V\'')
# title(ws, coordinates, offset=[1, 6], title='Average')
# title(ws, coordinates, offset=[2, 6], title='Std Dev')
# title(ws, coordinates, offset=[4, 6], title='SDCM')
# title(ws, coordinates, offset=[5, 6], title='Z score')
# title(ws, coordinates, offset=[6, 6], title='Failure Rate per 1000')
 # title(ws, coordinates, offset=[7, 6], title='Failure %')

wb.save(path)

# import os
# os.chdir("C://Users//mfevola//Documents//PythonScripts//PhotometricCompiler")
# import PhotometricCompiler